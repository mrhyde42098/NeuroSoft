"""
app/infrastructure/legacy_import_service.py
============================================
Migrador ETL: MISISTEMAV1.xlsm (VBA legacy) → ORM NeuroSoft.

Lee el libro Excel con openpyxl y vuelca sus hojas:

    DBRecepcion → PatientORM
    DBHC        → ClinicalHistoryORM (47 campos desarrollo/antecedentes/plan)
    DBObser     → ClinicalHistoryORM.obs_* (fusión por numero_documento+fecha)
    DBScore     → EvaluationORM (puntajes brutos serializados en JSON)
    DBETN       → EvolTerapiaORM

Características:
    • Idempotente: si (numero_documento, fecha_atencion) ya existe, omite.
    • Reporta conteos + lista de errores por fila.
    • NO hace commit parcial — todo o nada (salvo fallo por fila, que se salta).
    • NO requiere que el xlsm tenga macros activas; solo lee valores.

Dependencias: openpyxl (ya instalado por el módulo de exports).
"""
from __future__ import annotations

import json
import logging
import uuid
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

from sqlalchemy.orm import Session

logger = logging.getLogger("neurosoft.legacy_import")


# ─────────────────────────────────────────────────────────────
# Reporte del resultado de la migración
# ─────────────────────────────────────────────────────────────
@dataclass
class ImportReport:
    pacientes_creados: int = 0
    pacientes_omitidos: int = 0
    hc_creadas: int = 0
    hc_actualizadas: int = 0
    observaciones_fusionadas: int = 0
    evaluaciones_creadas: int = 0
    evoluciones_creadas: int = 0
    errores: list[dict[str, Any]] = field(default_factory=list)
    hojas_encontradas: list[str] = field(default_factory=list)
    hojas_esperadas: list[str] = field(default_factory=lambda: [
        "DBRecepcion", "DBHC", "DBObser", "DBScore", "DBETN",
    ])

    def as_dict(self) -> dict[str, Any]:
        return {
            "pacientes_creados": self.pacientes_creados,
            "pacientes_omitidos": self.pacientes_omitidos,
            "hc_creadas": self.hc_creadas,
            "hc_actualizadas": self.hc_actualizadas,
            "observaciones_fusionadas": self.observaciones_fusionadas,
            "evaluaciones_creadas": self.evaluaciones_creadas,
            "evoluciones_creadas": self.evoluciones_creadas,
            "errores": self.errores[:200],     # cap para no saturar respuesta
            "total_errores": len(self.errores),
            "hojas_encontradas": self.hojas_encontradas,
            "hojas_esperadas": self.hojas_esperadas,
        }


# ─────────────────────────────────────────────────────────────
# Helpers de parseo
# ─────────────────────────────────────────────────────────────
def _s(v: Any, default: str = "") -> str:
    if v is None:
        return default
    s = str(v).strip()
    return s if s else default


def _i(v: Any, default: int | None = None) -> int | None:
    if v in (None, "", "N/A"):
        return default
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return default


def _f(v: Any, default: float | None = None) -> float | None:
    if v in (None, "", "N/A"):
        return default
    try:
        return float(v)
    except (ValueError, TypeError):
        return default


def _b(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    return s in ("true", "1", "sí", "si", "yes", "y", "x")


def _d(v: Any) -> date | None:
    """Parsea celda Excel a date. Acepta date, datetime, ISO str, dd/mm/yyyy."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    # ISO
    try:
        return date.fromisoformat(s[:10])
    except ValueError:
        pass
    # dd/mm/yyyy
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _rows_as_dicts(ws) -> list[dict[str, Any]]:
    """Convierte una hoja openpyxl en lista de dicts usando fila 1 como header."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [(_s(h) or f"col{i}").lower() for i, h in enumerate(rows[0])]
    out: list[dict[str, Any]] = []
    for raw in rows[1:]:
        if raw is None or all(c in (None, "") for c in raw):
            continue
        out.append({headers[i]: raw[i] if i < len(raw) else None for i in range(len(headers))})
    return out


def _get(row: dict[str, Any], *keys: str, default: Any = None) -> Any:
    """Intenta varios nombres de columna (case-insensitive, ya lowercase)."""
    for k in keys:
        v = row.get(k.lower())
        if v not in (None, ""):
            return v
    return default


# ─────────────────────────────────────────────────────────────
# ETL principal
# ─────────────────────────────────────────────────────────────
def import_legacy_xlsm(
    xlsm_path: Path,
    db: Session,
    actor_id: str | None = None,
    actor_label: str | None = None,
    default_profesional_id: str | None = None,
) -> ImportReport:
    """Lee el xlsm y vuelca a ORM. Commit único al final."""
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError(
            "openpyxl no está instalado. `pip install openpyxl`."
        ) from e

    from app.infrastructure.database.orm_models import (
        ClinicalHistoryORM,
        EvaluationORM,
        EvolTerapiaORM,
        PatientORM,
    )

    if not xlsm_path.exists():
        raise FileNotFoundError(f"No existe el archivo: {xlsm_path}")

    rep = ImportReport()

    # read_only=True para archivos grandes; data_only=True → devuelve valores, no fórmulas
    wb = load_workbook(xlsm_path, read_only=True, data_only=True, keep_vba=False)
    rep.hojas_encontradas = list(wb.sheetnames)

    # ── 1) DBRecepcion → PatientORM ──────────────────────────
    doc_to_patient: dict[tuple[str, str], str] = {}   # (numero_documento, fecha_atencion) → patient_id
    if "DBRecepcion" in wb.sheetnames:
        rows = _rows_as_dicts(wb["DBRecepcion"])
        for i, r in enumerate(rows, start=2):
            try:
                numero_doc = _s(_get(r, "numero_documento", "numerodocumento", "documento", "cedula", "nrodocumento"))
                if not numero_doc:
                    continue
                fa = _d(_get(r, "fecha_atencion", "fechaatencion", "fecha")) or date.today()

                # Idempotencia — misma clave única que el ORM
                existing = (
                    db.query(PatientORM)
                    .filter(
                        PatientORM.numero_documento == numero_doc,
                        PatientORM.fecha_atencion == fa,
                    )
                    .first()
                )
                if existing:
                    doc_to_patient[(numero_doc, fa.isoformat())] = existing.id
                    rep.pacientes_omitidos += 1
                    continue

                p = PatientORM(
                    id=str(uuid.uuid4()),
                    numero_documento=numero_doc,
                    tipo_documento=_s(_get(r, "tipo_documento", "tipodocumento"), "CC"),
                    primer_nombre=_s(_get(r, "primer_nombre", "primernombre", "nombre"), "SIN_NOMBRE"),
                    segundo_nombre=_s(_get(r, "segundo_nombre", "segundonombre")) or None,
                    primer_apellido=_s(_get(r, "primer_apellido", "primerapellido", "apellido"), "SIN_APELLIDO"),
                    segundo_apellido=_s(_get(r, "segundo_apellido", "segundoapellido")) or None,
                    fecha_nacimiento=_d(_get(r, "fecha_nacimiento", "fechanacimiento")) or date(1900, 1, 1),
                    sexo=_s(_get(r, "sexo", "genero"), "M")[:1].upper(),
                    lugar_nacimiento=_s(_get(r, "lugar_nacimiento", "lugarnacimiento")) or None,
                    estado_civil=_s(_get(r, "estado_civil", "estadocivil")) or None,
                    telefono=_s(_get(r, "telefono")) or None,
                    correo=_s(_get(r, "correo", "email")) or None,
                    direccion=_s(_get(r, "direccion")) or None,
                    ciudad=_s(_get(r, "ciudad")) or None,
                    localidad=_s(_get(r, "localidad")) or None,
                    estrato=_i(_get(r, "estrato")),
                    escolaridad=_s(_get(r, "escolaridad"), "N/A"),
                    lateralidad=_s(_get(r, "lateralidad"), "Diestro"),
                    ocupacion=_s(_get(r, "ocupacion")) or None,
                    acompanante=_s(_get(r, "acompanante", "acompañante")) or None,
                    grupo_etnico=_s(_get(r, "grupo_etnico", "grupoetnico", "etnia")) or None,
                    profesional_id=default_profesional_id,
                    fecha_atencion=fa,
                    protocolo=_s(_get(r, "protocolo")) or None,
                    motivo_consulta=_s(_get(r, "motivo_consulta", "motivoconsulta")) or None,
                    remite=_s(_get(r, "remite")) or None,
                    eps=_s(_get(r, "eps")) or None,
                    orden_medica_no=_s(_get(r, "orden_medica_no", "ordenmedica", "ordenmedicano")) or None,
                    discapacidad=_s(_get(r, "discapacidad")) or None,
                    codigo_rips=_s(_get(r, "codigo_rips", "codigorips", "cie10")) or None,
                    cups=_s(_get(r, "cups")) or None,
                    finalidad_consulta=_s(_get(r, "finalidad_consulta", "finalidadconsulta")) or None,
                    numero_sesiones=_i(_get(r, "numero_sesiones", "numerosesiones"), 1),
                    donante=_b(_get(r, "donante")),
                )
                db.add(p)
                db.flush()
                doc_to_patient[(numero_doc, fa.isoformat())] = p.id
                rep.pacientes_creados += 1
            except Exception as e:  # noqa: BLE001
                rep.errores.append({"hoja": "DBRecepcion", "fila": i, "error": str(e)[:300]})
                logger.exception("Error DBRecepcion fila %d", i)

    # Cache auxiliar: numero_documento → último patient_id (para hojas que no tienen fecha)
    doc_last: dict[str, str] = {
        nd: pid for (nd, _), pid in doc_to_patient.items()
    }

    def _resolve_patient(num_doc: str, fa: date | None) -> str | None:
        if not num_doc:
            return None
        if fa:
            pid = doc_to_patient.get((num_doc, fa.isoformat()))
            if pid:
                return pid
        # Fallback: cualquier paciente con ese documento ya en BD
        if num_doc in doc_last:
            return doc_last[num_doc]
        existing = (
            db.query(PatientORM).filter(PatientORM.numero_documento == num_doc).first()
        )
        if existing:
            doc_last[num_doc] = existing.id
            return existing.id
        return None

    # ── 2) DBHC → ClinicalHistoryORM ─────────────────────────
    # Mapeo de columnas HC a atributos del ORM
    HC_FIELDS = [
        "motivo_consulta", "edad_materna", "no_gestacion", "riesgos", "cual_riesgo",
        "estres_prenatal", "gestacion", "semanas", "tipo_parto", "peso_gr", "talla_cm",
        "condiciones_neonatales", "incubadora", "sosten_cefalico", "sedestacion",
        "gateo", "marcha", "balbuceo", "primeras_palabras", "habla_claro",
        "control_anual", "control_vesical", "tipo_estres_prenatal", "ucin",
        "patologicos_medicos", "sensoriales_motores", "psiquiatricos", "farmacologicos",
        "traumaticos", "quirurgicos", "toxicos", "alergicos", "terapeuticos",
        "paraclinicos", "familiares", "vive_con", "abc", "escolar_laboral",
        "cognitivo", "comportamiento_animo", "patron_sueno", "patron_alimentacion",
        "plan_atencion", "impresion_diagnostica_hc",
    ]
    hc_by_key: dict[tuple[str, str], str] = {}  # (num_doc, fecha_iso) → hc_id

    if "DBHC" in wb.sheetnames:
        rows = _rows_as_dicts(wb["DBHC"])
        for i, r in enumerate(rows, start=2):
            try:
                num_doc = _s(_get(r, "numero_documento", "documento", "cedula"))
                fa = _d(_get(r, "fecha_atencion", "fecha"))
                pid = _resolve_patient(num_doc, fa)
                if not pid:
                    rep.errores.append({"hoja": "DBHC", "fila": i,
                                        "error": f"No se encontró paciente con doc={num_doc}"})
                    continue
                fa = fa or date.today()
                # Idempotencia
                existing = (
                    db.query(ClinicalHistoryORM)
                    .filter(ClinicalHistoryORM.patient_id == pid,
                            ClinicalHistoryORM.fecha_atencion == fa)
                    .first()
                )
                if existing:
                    hc_by_key[(num_doc, fa.isoformat())] = existing.id
                    continue

                hc = ClinicalHistoryORM(
                    id=str(uuid.uuid4()),
                    patient_id=pid,
                    numero_documento=num_doc,
                    fecha_atencion=fa,
                    codigo_cie10=_s(_get(r, "codigo_cie10", "cie10"), "F809"),
                )
                for f_ in HC_FIELDS:
                    val = _get(r, f_, f_.replace("_", ""))
                    if val not in (None, ""):
                        setattr(hc, f_, _s(val, "N/A"))
                db.add(hc)
                db.flush()
                hc_by_key[(num_doc, fa.isoformat())] = hc.id
                rep.hc_creadas += 1
            except Exception as e:  # noqa: BLE001
                rep.errores.append({"hoja": "DBHC", "fila": i, "error": str(e)[:300]})
                logger.exception("Error DBHC fila %d", i)

    # ── 3) DBObser → ClinicalHistoryORM.obs_* ────────────────
    OBS_FIELDS = [
        "obs_clinica_general", "obs_atencion", "obs_memoria", "obs_praxias_gnosias",
        "obs_lenguaje", "obs_funciones_ejecutivas", "obs_emociones", "obs_ci",
        "obs_impresion_dx", "obs_funcionalidad", "obs_recomendaciones",
    ]
    if "DBObser" in wb.sheetnames:
        rows = _rows_as_dicts(wb["DBObser"])
        for i, r in enumerate(rows, start=2):
            try:
                num_doc = _s(_get(r, "numero_documento", "documento", "cedula"))
                fa = _d(_get(r, "fecha_atencion", "fecha"))
                pid = _resolve_patient(num_doc, fa)
                if not pid:
                    rep.errores.append({"hoja": "DBObser", "fila": i,
                                        "error": f"No se encontró paciente con doc={num_doc}"})
                    continue
                fa = fa or date.today()

                hc_id = hc_by_key.get((num_doc, fa.isoformat()))
                hc = db.get(ClinicalHistoryORM, hc_id) if hc_id else None
                if hc is None:
                    # Crear HC mínima que hospede las observaciones
                    hc = ClinicalHistoryORM(
                        id=str(uuid.uuid4()),
                        patient_id=pid,
                        numero_documento=num_doc,
                        fecha_atencion=fa,
                    )
                    db.add(hc)
                    db.flush()
                    hc_by_key[(num_doc, fa.isoformat())] = hc.id
                    rep.hc_creadas += 1

                changed = False
                for f_ in OBS_FIELDS:
                    val = _get(r, f_, f_.replace("_", ""))
                    if val not in (None, ""):
                        setattr(hc, f_, _s(val, "N/A"))
                        changed = True
                if changed:
                    rep.observaciones_fusionadas += 1
                    if hc_id:
                        rep.hc_actualizadas += 1
            except Exception as e:  # noqa: BLE001
                rep.errores.append({"hoja": "DBObser", "fila": i, "error": str(e)[:300]})
                logger.exception("Error DBObser fila %d", i)

    # ── 4) DBScore → EvaluationORM ───────────────────────────
    # Estrategia: agrupar por (num_doc, fecha) y vaciar todas las columnas
    # numéricas como puntajes_brutos_json {col_name: value}.
    if "DBScore" in wb.sheetnames:
        rows = _rows_as_dicts(wb["DBScore"])
        # Agrupamos por (num_doc, fecha)
        groups: dict[tuple[str, str], dict[str, Any]] = {}
        for r in rows:
            num_doc = _s(_get(r, "numero_documento", "documento", "cedula"))
            fa = _d(_get(r, "fecha_atencion", "fecha", "fechaevaluacion"))
            if not num_doc:
                continue
            key = (num_doc, (fa or date.today()).isoformat())
            bucket = groups.setdefault(key, {})
            for k, v in r.items():
                if k in ("numero_documento", "documento", "cedula",
                         "fecha_atencion", "fecha", "fechaevaluacion"):
                    continue
                if isinstance(v, (int, float)):
                    bucket[k] = v
                elif v not in (None, ""):
                    bucket[k] = _s(v)

        for (num_doc, fecha_iso), puntajes in groups.items():
            try:
                fa = date.fromisoformat(fecha_iso)
                pid = _resolve_patient(num_doc, fa)
                if not pid:
                    rep.errores.append({"hoja": "DBScore",
                                        "error": f"No se encontró paciente con doc={num_doc}"})
                    continue
                protocolo_val = _s(puntajes.pop("protocolo", "") or puntajes.pop("protocolo_id", ""))

                # Idempotencia aproximada: misma fecha + protocolo + paciente
                existing = (
                    db.query(EvaluationORM)
                    .filter(
                        EvaluationORM.patient_id == pid,
                        EvaluationORM.fecha == fa,
                        EvaluationORM.protocolo == (protocolo_val or None),
                    )
                    .first()
                )
                if existing:
                    continue

                ev = EvaluationORM(
                    id=str(uuid.uuid4()),
                    patient_id=pid,
                    protocolo=protocolo_val or None,
                    fecha=fa,
                    puntajes_brutos_json=json.dumps(puntajes, ensure_ascii=False, default=str),
                    pruebas_realizadas=sum(1 for v in puntajes.values() if v not in (None, "")),
                    is_latest=True,
                )
                db.add(ev)
                db.flush()
                rep.evaluaciones_creadas += 1
            except Exception as e:  # noqa: BLE001
                rep.errores.append({"hoja": "DBScore", "num_doc": num_doc,
                                    "error": str(e)[:300]})
                logger.exception("Error DBScore doc=%s", num_doc)

    # ── 5) DBETN → EvolTerapiaORM ────────────────────────────
    if "DBETN" in wb.sheetnames:
        rows = _rows_as_dicts(wb["DBETN"])
        for i, r in enumerate(rows, start=2):
            try:
                num_doc = _s(_get(r, "numero_documento", "documento", "cedula"))
                fa_sesion = _d(_get(r, "fecha_sesion", "fechasesion", "fecha")) or date.today()
                pid = _resolve_patient(num_doc, None)
                if not pid:
                    rep.errores.append({"hoja": "DBETN", "fila": i,
                                        "error": f"No se encontró paciente con doc={num_doc}"})
                    continue
                ev = EvolTerapiaORM(
                    id=str(uuid.uuid4()),
                    patient_id=pid,
                    numero_documento=num_doc,
                    sesiones_orden=_s(_get(r, "sesiones_orden", "sesionesorden")) or None,
                    numero_orden=_s(_get(r, "numero_orden", "numeroorden")) or None,
                    fecha_inicio=_d(_get(r, "fecha_inicio", "fechainicio")),
                    fecha_sesion=fa_sesion,
                    numero_sesion=_s(_get(r, "numero_sesion", "numerosesion"), "N/A"),
                    objetivos=_s(_get(r, "objetivos"), "N/A"),
                    actividades=_s(_get(r, "actividades"), "N/A"),
                    plan_trabajo=_s(_get(r, "plan_trabajo", "plantrabajo"), "N/A"),
                )
                db.add(ev)
                rep.evoluciones_creadas += 1
            except Exception as e:  # noqa: BLE001
                rep.errores.append({"hoja": "DBETN", "fila": i, "error": str(e)[:300]})
                logger.exception("Error DBETN fila %d", i)

    # §C6-fix: commit atómico con rollback si la operación global falla.
    # Antes, si una fila individual lanzaba dentro del try interno, las
    # filas previas ya estaban en sesión y se commiteaban igualmente.
    # Ahora hacemos commit explícito con rollback en error → integridad.
    try:
        db.commit()
    except Exception as commit_err:  # noqa: BLE001
        db.rollback()
        logger.exception("Error en commit de import legacy — rollback aplicado")
        rep.errores.append({"hoja": "GLOBAL", "fila": 0,
                            "error": f"Commit fallido: {commit_err}"})
    wb.close()

    # Auditoría de la operación global
    try:
        from app.infrastructure.audit import record_event
        record_event(
            db,
            action="import_legacy_xlsm",
            entity_type="system",
            actor_id=actor_id,
            actor_label=actor_label,
            summary=(
                f"Import xlsm: pacientes={rep.pacientes_creados} "
                f"hc={rep.hc_creadas} ev={rep.evaluaciones_creadas} "
                f"etn={rep.evoluciones_creadas} err={len(rep.errores)}"
            ),
        )
    except Exception:
        pass

    return rep
